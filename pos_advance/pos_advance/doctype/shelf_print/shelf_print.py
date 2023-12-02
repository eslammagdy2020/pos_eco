# Copyright (c) 2023, beshoyatef31@gmail.com and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe import _
import openpyxl
from pos_advance.utils import scan_barcode

class ShelfPrint(Document):
	
	def validate_price_list(self):
		selling = frappe.db.get_value('Price List', self.price_list , 'selling')
		if not selling :
			frappe.throw(f"Price list {selling} is not Selling Price List ")
	def validate(self):
		self.validate_price_list()
		self.get_manuale_values()
		if self.sheet_date :
			self.get_data_from_file()
		self.set_item_price()
		self.validate_price_rule()


	def get_manuale_values(self) :
		if not self.sheet_date :
			if self.items and len(self.items) > 0 :
				for row in self.items :
					if not row.code :
						frappe.thorw(f"Please set Code in raw {row}")
					item = scan_barcode(row.code)
					row.item =item.get("item_code")
					row.item_name  =  frappe.db.get_value("Item" , 
									{"item_code" : row.item  } ,"item_name" )	
					row.uom = item.get("uom")


	def get_data_from_file(self):
		if not self.sheet_date :
			frappe.throw(_("Please Select excel sheet !"))
		
		pat = f"{self.sheet_date}".split('/')
		workbook =  openpyxl.load_workbook(frappe.get_site_path('private', 'files', pat[-1]) )
		sheet_obj = workbook.active
		self.items = []
		for i in range(2, int(sheet_obj.max_row) + 1 ) :
			cell_obj = sheet_obj.cell(row = i, column = 1)
			item = scan_barcode(cell_obj.value)
			if not item :
				frappe.msgprint(_(f"Can not find Barcode {cell_obj.value} "))
			if item :
				row = self.append("items" , {})
				row.item =item.get("item_code")
				row.item_name  =  frappe.db.get_value("Item" , 
				{"item_code" : row.item  } ,"item_name" )	
				row.uom = item.get("uom")
				row.code = cell_obj.value
			# print(workbook)

	def validate_price_rule(self) :
		if self.price_rule :
			if self.price_list != frappe.get_value("Pricing Rule" ,self.price_rule , "for_price_list") :
				frappe.throw(_( f"""You Set Price List {self.price_list} but Price Rule {self.price_rule} is work for 
				 {frappe.get_value("Pricing Rule" ,self.price_rule , "for_price_list") } """ ))
				return 0
			# if price rule apply on  Apply On  tansaction will throw invalid
			if frappe.get_value("Pricing Rule" ,self.price_rule , "apply_on") == "Transaction" :
					frappe.throw(_(""" Not Allowed To print Shelf Fro transaction """))

	def get_item_to_apply_prcie_rule(self):
		if self.price_rule :
			# frappe.cache().set_value(f"erpnext:barcode_scan:{search_value}", data, expires_in_sec=120)	
		    # get all items rice rule applicable for 
			data  = frappe.cache().get_value(f"posadvanc_{self.price_rule}")
			if data :
				print(f"Get Cahched Data , {data}" )
				return data 
			if not data :
				items_codes = frappe.db.sql(f""" 
				select name FROM `tabItem` WHERE item_group in 
				(SELECT item_group FROM `tabPricing Rule Item Group` WHERE parent ='{self.price_rule}') 
				OR 
				brand in 
				(SELECT brand FROM `tabPricing Rule Brand` WHERE parent = '{self.price_rule}') 
				OR 
				name in 
				(SELECT item_code from `tabPricing Rule Item Code` WHERE parent= '{self.price_rule}')
				
				""")
				
				if len(items_codes ) > 0 :
					data = (i[0] for i in items_codes)
					frappe.cache().set_value(f"posadvanc_{self.price_rule}", 
					  [i[0] for i in items_codes ] , expires_in_sec= 120 )
				
			return data 

	def caculate_discount(self , item_price) :
		price = float(item_price or 0 )
		if price <=  0 :
			frappe.msgprint(" Price is Zero amount ")
			return price 
		rule = frappe.get_doc("Pricing Rule" ,self.price_rule)
		if rule.rate_or_discount == "Rate":
			price = rule.rate
		if rule.rate_or_discount == "Discount Percentage":
			price = float(item_price or 0 )  - ((float(rule.discount_percentage or 0)  / 100 ) 
							* float(item_price or 0 ))
		if rule.rate_or_discount == "Discount Amount ":
			price = float(item_price or  0 ) - float(rule.discount_amount or 0 )

		return price
	def set_item_price(self):
		if self.items and len(self.items) > 0 :
		
			for item in self.items :
				item.price = frappe.db.get_value("Item Price" ,
				  {"price_list" :self.price_list , "uom" :item.uom , "item_code" : item.item}
				  ,"price_list_rate" )	
				item.price_after_discount = item.price
				item.barcode = item.code 
				if self.price_rule : 
					#caculate_discount 
					#check if item has price Rule 
					codes = self.get_item_to_apply_prcie_rule() or []
					if item.item in codes :
						item.has_discount = 1
						item.price_after_discount = self.caculate_discount(item.price)
				if float(item.price or 0 ) ==  0 :
					frappe.msgprint(f""" Can not find item price for item {item.item} \
							in price list {self.price_list} for uom {item.uom}""")
					# self.items.remove(item.idx)
